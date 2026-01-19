import streamlit as st
import pandas as pd
from datetime import datetime
import os
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

EXCEL_PATH = "job_applications.xlsx"

if os.path.exists(EXCEL_PATH):
    df = pd.read_excel(EXCEL_PATH, engine="openpyxl")
else:
    df = pd.DataFrame(
        columns=["Company Name", "JD URL", "Applied Time", "Heard From Them?"]
    )

st.title("📋 Job Application Tracker")

with st.form("job_form", clear_on_submit=True):
    company = st.text_input("Company Name")
    url = st.text_input("Job Posting URL")
    heard = st.checkbox("Heard from them?")
    submitted = st.form_submit_button("Save")

if submitted and company and url:
    now = datetime.now()

    new_row = {
        "Company Name": company,
        "JD URL": url,
        "Applied Time": now,
        "Heard From Them?": "Yes" if heard else "No"
    }

    df = pd.concat([df, pd.DataFrame([new_row])], ignore_index=True)
    df.to_excel(EXCEL_PATH, index=False, engine="openpyxl")

    wb = load_workbook(EXCEL_PATH)
    ws = wb.active

    header_font = Font(bold=True)
    center_align = Alignment(horizontal="center", vertical="center")

    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.alignment = center_align
        ws.column_dimensions[get_column_letter(col)].width = 28

    ws.freeze_panes = "A2"

    # Hyperlinks + datetime format
    for row in range(2, ws.max_row + 1):
        url_cell = ws.cell(row=row, column=2)
        if url_cell.value:
            url_cell.hyperlink = url_cell.value
            url_cell.font = Font(color="0000FF", underline="single")

        date_cell = ws.cell(row=row, column=3)
        date_cell.number_format = "yyyy-mm-dd hh:mm"

    wb.save(EXCEL_PATH)

    st.success("✅ Application saved!")
